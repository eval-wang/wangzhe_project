# -*- coding: utf-8 -*-
"""
子项目3_数据定位 —— 基础表格填充脚本

功能：
    1. 读取 hero_grid_result.csv 中的采样数据 (file, time_s, row, col, status)，
       用 (row, col) 查"小地图位置"映射表现算符号，
       按 15 秒一个格子的时间网格填入基础表格，另存为结果文件（原件不动）。
    2. 填充表格生成后，在输出目录下创建 UN/ 和 EM/ 两个文件夹：
       - UN/：识别到英雄但 (row, col) 落在小地图未标注格子的采样
       - EM/：status=empty（小地图未识别到英雄）的采样
       各文件夹内生成一份 CSV（出现的文件名称和次数），
       并将配套截图（与 CSV 同目录的 frame 图片）复制进去。

规则（详见 预期外事件/项目3/项目3 copy.md）：
    - 两个 CSV 自带的"符号"列不使用，仅以 (row, col) 查映射表。
    - time_s < 40 的采样点丢弃（表格从 0:40 开始）。
    - 格子为 15 秒时间窗，左闭右开；恰好落在边界的点归入后一格。
    - status=empty -> "EM"；识别到但查不到映射 -> "UN"。
    - 同一格内多个事件按时间顺序用 "_" 拼接，如 "b0_b1"、"EM_b3"。
    - 15:40 之后的格子无数据，自然留空。
"""

import csv
import shutil
from collections import Counter
from pathlib import Path

import openpyxl

# ---------------- 可配置路径 ----------------
CSV_PATH = Path(r"D:\wangzhe_project\screenshot_collection\test_collection\test1\hero_grid_result.csv")
MAP_PATH = Path(r"D:\wangzhe_project\王者视频解析\子项目3_数据定位\小地图位置.md")
BASE_TABLE_PATH = Path(r"D:\wangzhe_project\相关素材\表格\基础表格.xlsx")
# 产物目录命名约定（可追踪、可复盘）：产物根目录 + 阶段名，
# 阶段文件夹由脚本自动创建，最终产物放在 <APPLY_ROOT>/<STAGE_NAME>/ 下。
APPLY_ROOT = Path(r"D:\wangzhe_project\apply\document1")
STAGE_NAME = "subproject3"
OUTPUT_DIR = APPLY_ROOT / STAGE_NAME
OUTPUT_PATH = OUTPUT_DIR / "基础表格_填充结果.xlsx"
# 配套截图所在目录（默认与采样 CSV 同目录）
IMAGE_DIR = CSV_PATH.parent

# ---------------- 网格参数 ----------------
GRID_START_S = 40        # 表格起点 0:40（秒）
CELL_SPAN_S = 15         # 每格 15 秒
ROWS_PER_COL = 6         # 每列 6 行（6 × 15s = 90s = 一个列宽）
EMPTY_MARK = "EM"        # empty 帧（未识别到英雄）
UNMAPPED_MARK = "UN"     # 识别到但落在未标注格子
SEP = "_"                # 同格多事件拼接符


def load_symbol_map(map_path: Path) -> dict:
    """读取小地图位置映射表：(行, 列) -> 符号。"""
    symbol_map = {}
    with open(map_path, encoding="utf-8-sig", newline="") as f:
        for record in csv.DictReader(f):
            key = (int(record["行"]), int(record["列"]))
            symbol_map[key] = record["符号"].strip()
    return symbol_map


def load_samples(csv_path: Path) -> list:
    """读取采样数据，返回 [(file, time_s, row, col, status), ...]，按时间排序。"""
    samples = []
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        for record in csv.DictReader(f):
            time_s = float(record["time_s"])
            status = record["status"].strip()
            row = int(record["row"]) if record["row"].strip() else None
            col = int(record["col"]) if record["col"].strip() else None
            samples.append((record["file"].strip(), time_s, row, col, status))
    samples.sort(key=lambda s: s[1])
    return samples


def classify_sample(sample, symbol_map) -> str:
    """单个采样点 -> 格子事件文本（符号 / EM / UN）。"""
    _file, _time_s, row, col, status = sample
    if status != "found" or row is None or col is None:
        return EMPTY_MARK
    return symbol_map.get((row, col), UNMAPPED_MARK)


def build_grid(samples, symbol_map, n_cols: int):
    """
    将采样点分配到时间格子。
    返回 (grid, placed)：
      grid   — {(col_idx, row_idx): [事件文本, ...]}，下标从 0 开始；
      placed — 实际落入表格范围内的采样点列表（供 UN/EM 统计）。
    格子 [col_idx, row_idx] 覆盖 [GRID_START_S + col_idx*90 + row_idx*15, +15s)，左闭右开。
    """
    grid = {}
    placed = []
    for sample in samples:
        time_s = sample[1]
        if time_s < GRID_START_S:
            continue  # 前 40 秒丢弃
        offset = time_s - GRID_START_S
        col_idx = int(offset // (CELL_SPAN_S * ROWS_PER_COL))
        row_idx = int((offset % (CELL_SPAN_S * ROWS_PER_COL)) // CELL_SPAN_S)
        if col_idx >= n_cols:
            continue  # 超出表格范围
        grid.setdefault((col_idx, row_idx), []).append(classify_sample(sample, symbol_map))
        placed.append(sample)
    return grid, placed


def export_event_folder(kind: str, samples, symbol_map):
    """
    创建 UN/ 或 EM/ 文件夹，写入事件 CSV 并复制配套截图。
    kind: "UN" 或 "EM"。返回事件总数。
    """
    folder = OUTPUT_DIR / kind
    if folder.exists():
        shutil.rmtree(folder)  # 清除上次运行的结果，避免残留旧图
    folder.mkdir(parents=True)

    # 筛选目标采样点
    if kind == UNMAPPED_MARK:
        targets = [s for s in samples
                   if s[4] == "found" and s[2] is not None and (s[2], s[3]) not in symbol_map]
        header = ["file", "time_s", "row", "col", "count"]
        key_of = lambda s: (s[0], s[2], s[3])          # 按 文件+位置 聚合
        row_of = lambda key, cnt: [key[0], next(s[1] for s in targets if s[0] == key[0]),
                                   key[1], key[2], cnt]
    else:
        targets = [s for s in samples if s[4] != "found" or s[2] is None]
        header = ["file", "time_s", "count"]
        key_of = lambda s: (s[0],)                      # 按文件聚合
        row_of = lambda key, cnt: [key[0], next(s[1] for s in targets if s[0] == key[0]), cnt]

    counter = Counter(key_of(s) for s in targets)

    csv_file = folder / f"{kind}.csv"
    with open(csv_file, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for key, cnt in sorted(counter.items(), key=lambda kv: kv[0][0]):
            writer.writerow(row_of(key, cnt))

    # 复制配套截图（存在才复制）
    copied, missing = 0, []
    for file_name in {s[0] for s in targets}:
        src = IMAGE_DIR / file_name
        if src.exists():
            shutil.copy2(src, folder / file_name)
            copied += 1
        else:
            missing.append(file_name)

    print(f"[{kind}] 事件 {len(targets)} 次，涉及文件 {len(counter)} 个，"
          f"已复制图片 {copied} 张" + (f"，缺失图片: {missing}" if missing else ""))
    return len(targets)


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)  # 自动创建阶段文件夹
    symbol_map = load_symbol_map(MAP_PATH)
    samples = load_samples(CSV_PATH)
    print(f"映射表条目: {len(symbol_map)}，采样点: {len(samples)}")

    wb = openpyxl.load_workbook(BASE_TABLE_PATH)
    ws = wb["Sheet1"]

    # 列头在第 1 行（B 列起），行标签在 A 列（第 2 行起），行列数以表头实际内容为准
    col_headers = [cell.value for cell in ws[1][1:] if cell.value is not None]
    row_labels = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    n_cols = len(col_headers)
    print(f"表格网格: {n_cols} 列 × {len(row_labels)} 行，列起点: {col_headers[0]} ~ {col_headers[-1]}")

    grid, placed = build_grid(samples, symbol_map, n_cols)

    filled = 0
    for (col_idx, row_idx), events in grid.items():
        cell = ws.cell(row=2 + row_idx, column=2 + col_idx)
        cell.value = SEP.join(events)
        cell.number_format = "@"  # 强制文本，避免被识别为时间/数字
        filled += 1

    wb.save(OUTPUT_PATH)
    print(f"填充格子: {filled}，结果已保存: {OUTPUT_PATH}")

    # 生成 UN / EM 文件夹（只统计实际落入表格的采样点）
    export_event_folder(UNMAPPED_MARK, placed, symbol_map)
    export_event_folder(EMPTY_MARK, placed, symbol_map)


if __name__ == "__main__":
    main()
