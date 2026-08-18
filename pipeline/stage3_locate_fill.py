# -*- coding: utf-8 -*-
"""
Pipeline 阶段3 —— 网格定位 + 符号标注 + 填充基础表格 + UN 归档
================================================================
消费阶段2 的 hero_position.csv（px, py），负责：
  1. 像素坐标 → 20×20 网格（floor(px / 9.3)，原点左上角，边界钳制）
  2. (row, col) → 查 小地图位置.md → 符号，生成 hero_grid_result.csv
  3. 按 15 秒时间格填充基础表格 → 基础表格_填充结果.xlsx（模板只读，另存）
  4. UN 归档：found 但查不到映射的帧 → UN\\UN.csv + 截图

填表规则（详见 pipeline串联_问答记录.md）：
  - 时间基准 40s 起，15 秒一格，每列 6 行（90 秒），左闭右开，边界点归入后一格
  - 单元格值：符号 / EM（empty 帧）/ UN（未标注点），同格多事件按时间顺序用 "_" 拼接，不去重
  - 超出基础表格列数（17 列，约 26:10）的采样丢弃

依赖（只读）：
  - 符号映射：王者视频解析\\子项目3_数据定位\\小地图位置.md
  - 表格模板：相关素材\\表格\\基础表格.xlsx
"""

import csv
import shutil
from collections import Counter
from pathlib import Path

import openpyxl

# ---------------- 只读引用 ----------------
SYMBOL_MAP_PATH = Path(r"D:\wangzhe_project\王者视频解析\子项目3_数据定位\小地图位置.md")
BASE_TABLE_PATH = Path(r"D:\wangzhe_project\相关素材\表格\基础表格.xlsx")

# ---------------- 网格与表格参数 ----------------
GRID_SIZE = 20           # 20×20 网格
CELL_PX = 9.3            # 186 / 20 = 9.3 px/格
GRID_START_S = 40        # 表格起点 0:40（秒）
CELL_SPAN_S = 15         # 每格 15 秒
ROWS_PER_COL = 6         # 每列 6 行（6 × 15s = 90s = 一个列宽）
EMPTY_MARK = "EM"        # empty 帧（未识别到英雄）
UNMAPPED_MARK = "UN"     # 识别到但落在未标注格子
SEP = "_"                # 同格多事件拼接符


def pixel_to_grid(px, py):
    """像素坐标 → 网格 (row, col)，floor 换算 + 边界钳制。"""
    col = max(0, min(GRID_SIZE - 1, int(px // CELL_PX)))
    row = max(0, min(GRID_SIZE - 1, int(py // CELL_PX)))
    return row, col


def load_symbol_map(map_path=SYMBOL_MAP_PATH):
    """读取 (行,列) → 符号 映射表。"""
    symbol_map = {}
    with open(map_path, encoding="utf-8-sig", newline="") as f:
        for record in csv.DictReader(f):
            symbol_map[(int(record["行"]), int(record["列"]))] = record["符号"].strip()
    return symbol_map


def load_positions(csv_path):
    """读取阶段2 的 hero_position.csv，返回 [(file, time_s, status, px, py), ...]。"""
    samples = []
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        for record in csv.DictReader(f):
            px = int(record["px"]) if record["px"].strip() else None
            py = int(record["py"]) if record["py"].strip() else None
            samples.append((record["file"].strip(), float(record["time_s"]),
                            record["status"].strip(), px, py))
    samples.sort(key=lambda s: s[1])
    return samples


def locate_and_fill(position_csv, frames_dir, output_dir,
                    symbol_map_path=SYMBOL_MAP_PATH, base_table_path=BASE_TABLE_PATH):
    """
    阶段3 主流程。

    Args:
        position_csv: 阶段2 的 hero_position.csv
        frames_dir: 阶段1 的小地图帧目录（用于 UN 截图归档）
        output_dir: 阶段3 产物目录（<APPLY_ROOT>\\<document>\\subproject3），自动创建
        symbol_map_path: 符号映射表
        base_table_path: 基础表格模板（只读）

    Returns:
        dict: 各项产物路径与统计
    """
    position_csv = Path(position_csv)
    frames_dir = Path(frames_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)  # 阶段文件夹自动创建

    symbol_map = load_symbol_map(symbol_map_path)
    samples = load_positions(position_csv)
    print(f"[阶段3] 映射表条目: {len(symbol_map)}，采样点: {len(samples)}")

    # ---------- 1+2: 网格 + 符号 → hero_grid_result.csv ----------
    grid_csv = output_dir / "hero_grid_result.csv"
    events = []  # (file, time_s, event_text, row, col, status)
    with open(grid_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["file", "time_s", "row", "col", "status", "px", "py", "符号"])
        for file, time_s, status, px, py in samples:
            if status == "found" and px is not None:
                row, col = pixel_to_grid(px, py)
                sym = symbol_map.get((row, col), "")
                writer.writerow([file, time_s, row, col, status, px, py, sym])
                events.append((file, time_s, sym if sym else UNMAPPED_MARK, row, col, status))
            else:
                writer.writerow([file, time_s, "", "", "empty", "", "", ""])
                events.append((file, time_s, EMPTY_MARK, None, None, "empty"))

    # ---------- 3: 填充基础表格 ----------
    wb = openpyxl.load_workbook(base_table_path)
    ws = wb["Sheet1"]
    col_headers = [c.value for c in ws[1][1:] if c.value is not None]
    n_cols = len(col_headers)
    n_rows = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value is not None)

    grid = {}
    placed = []
    for file, time_s, event, row, col, status in events:
        if time_s < GRID_START_S:
            continue
        offset = time_s - GRID_START_S
        col_idx = int(offset // (CELL_SPAN_S * ROWS_PER_COL))
        row_idx = int((offset % (CELL_SPAN_S * ROWS_PER_COL)) // CELL_SPAN_S)
        if col_idx >= n_cols or row_idx >= n_rows:
            continue  # 超出表格范围丢弃
        grid.setdefault((col_idx, row_idx), []).append(event)
        placed.append((file, time_s, event, row, col, status))

    for (col_idx, row_idx), evs in grid.items():
        cell = ws.cell(row=2 + row_idx, column=2 + col_idx)
        cell.value = SEP.join(evs)
        cell.number_format = "@"

    table_path = output_dir / "基础表格_填充结果.xlsx"
    wb.save(table_path)
    print(f"[阶段3] 表格网格 {n_cols} 列 × {n_rows} 行，填充格子: {len(grid)} → {table_path.name}")

    # ---------- 4: UN 归档（只统计落入表格的采样，先清空再生成） ----------
    un_dir = output_dir / UNMAPPED_MARK
    if un_dir.exists():
        shutil.rmtree(un_dir)
    un_dir.mkdir(parents=True)

    un_items = [p for p in placed if p[2] == UNMAPPED_MARK]
    counter = Counter((p[0], p[3], p[4]) for p in un_items)
    with open(un_dir / f"{UNMAPPED_MARK}.csv", "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["file", "time_s", "row", "col", "count"])
        for (name, row, col), cnt in sorted(counter.items()):
            time_s = next(p[1] for p in un_items if p[0] == name)
            writer.writerow([name, time_s, row, col, cnt])

    copied = 0
    for name in {p[0] for p in un_items}:
        src = frames_dir / name
        if src.exists():
            shutil.copy2(src, un_dir / name)
            copied += 1

    print(f"[阶段3] UN 事件 {len(un_items)} 次，归档 {copied} 张截图 → {un_dir}")
    return {"grid_csv": grid_csv, "table_path": table_path, "un_dir": un_dir,
            "un_count": len(un_items), "filled_cells": len(grid)}
